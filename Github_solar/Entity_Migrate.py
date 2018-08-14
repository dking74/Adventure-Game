from openpyxl import Workbook
from orionsdk import SwisClient
import openpyxl, requests, sys
import numpy as np 

class IntelligridMig ( ):

    '''
        Class name    : IntelligridMig
    
        Class Purpose : To migrate entities from Excel to Solarwinds
    '''

    def __init__        ( self , inputFile ):

        '''
            Method name    : __init__
        
            Method Purpose : To initialize the SolarWinds migration instance
        
            Parameters     :
                -inputFile : The file to be loaded
        
            Returns        : None
        '''

        # the server is unverified --> allow without warnings
        verify = False
        if not verify:
            from requests.packages.urllib3.exceptions import InsecureRequestWarning
            requests.packages.urllib3.disable_warnings ( InsecureRequestWarning )

        # login to solarwinds and load the worksheet
        self.loginSolar     (           )
        self.load_worksheet ( inputFile )

        # get top level group
        testGroup   = self.createGroup ( "Test" , "This is the test, base-level group" , [] )
        testGroupID = self.getGroupID  ( "Test" )

        # get secondary level groups
        aicGroup         = self.createGroup ( "AIC" ,  "This is the test, base-level group" , [] )
        argGroup         = self.createGroup ( "ARG" ,  "This is the test, base-level group" , [] )
        aueGroup         = self.createGroup ( "AUE" ,  "This is the test, base-level group" , [] )
        uecGroup         = self.createGroup ( "UEC" ,  "This is the test, base-level group" , [] )
        ipcGroup         = self.createGroup ( "IPC" ,  "This is the test, base-level group" , [] )
        aepGroup         = self.createGroup ( "AEP" ,  "This is the test, base-level group" , [] )
        sweGroup         = self.createGroup ( "SWEC" , "This is the test, base-level group" , [] )
        aecGroup         = self.createGroup ( "AECI" , "This is the test, base-level group" , [] )
        itcGroup         = self.createGroup ( "ITC" ,  "This is the test, base-level group" , [] )
        cipGroup         = self.createGroup ( "CIP" ,  "This is the test, base-level group" , [] )
        unkGroup         = self.createGroup ( "UNK" ,  "This is the test, base-level group" , [] )

        # list of the sub-level group info
        midGroups = [ 
                        aicGroup + argGroup + aueGroup + uecGroup + ipcGroup + aepGroup + \
                        sweGroup + aecGroup + itcGroup + cipGroup + unkGroup
                    ]

        # add sub groups to main group
        self.addDefinitions ( testGroupID [ 'results' ][ 0 ][ 'ContainerID' ] , midGroups [ 0 ] )

    def loginSolar      ( self ):

        '''
            Method name    : loginSolar
        
            Method Purpose : To create an instance of solarwinds
        
            Parameters     : None
        
            Returns        : None
        '''

        valid = False        

        # get the username and password from the user until it is valid
        while not valid:

            username = input ( "Username: " )
            password = input ( "Password: " )
            print            (     "\n"     )

            try: 
                self._solarwinds = SwisClient ( "solarwinds.ameren.com" , username , password )
                self._solarwinds.query        (     "SELECT AccountID FROM Orion.Accounts"    )
                valid = True

            except requests.exceptions.HTTPError:
                valid = False

    def load_worksheet  ( self , file ):
    

        '''
            Method name    : load_workbook
        
            Method Purpose : To load the workbook/worsheet
        
            Parameters     : None
        
            Returns        : None
        '''

        # get the workbook and worksheet locaclly
        self._intelligridBook  = openpyxl.load_workbook ( file )
        self._intelligridSheet = self._intelligridBook.active

    def read_workbook   ( self ):

        '''
            Method name    : read_workbook
        
            Method Purpose : To read the inputted workbook
        
            Parameters     : None
        
            Returns        : None
        '''

        # read every row in the book
        for ROW in range ( 3 , self._intelligridSheet.max_row + 1 ):

            # get the column info from row
            legacy_loc = self._intelligridSheet.cell ( row=ROW , column=1  ).value
            prefix     = self._intelligridSheet.cell ( row=ROW , column=2  ).value
            action_req = self._intelligridSheet.cell ( row=ROW , column=3  ).value
            urgency    = self._intelligridSheet.cell ( row=ROW , column=4  ).value
            site_id    = self._intelligridSheet.cell ( row=ROW , column=5  ).value
            site_check = self._intelligridSheet.cell ( row=ROW , column=6  ).value
            loc_name   = self._intelligridSheet.cell ( row=ROW , column=7  ).value
            division   = self._intelligridSheet.cell ( row=ROW , column=8  ).value
            owning_co  = self._intelligridSheet.cell ( row=ROW , column=9  ).value
            asset_type = self._intelligridSheet.cell ( row=ROW , column=10 ).value
            latitude   = self._intelligridSheet.cell ( row=ROW , column=11 ).value
            longitude  = self._intelligridSheet.cell ( row=ROW , column=12 ).value
            address    = self._intelligridSheet.cell ( row=ROW , column=13 ).value
            loc_id     = self._intelligridSheet.cell ( row=ROW , column=14 ).value
            emprv_dist = self._intelligridSheet.cell ( row=ROW , column=15 ).value
            prim_dist  = self._intelligridSheet.cell ( row=ROW , column=16 ).value


            # find if there is any nodes matching legacy code
            leg_search = {}
            if legacy_loc != '':
                leg_search  = self._solarwinds.query (  """
                                                        SELECT 
                                                            Caption,
                                                            Uri
                                                        FROM 
                                                            Orion.Nodes 
                                                        WHERE
                                                            Caption
                                                        LIKE '{}%'
                                                        """.format ( str ( legacy_loc ).lower ( ) ) )                                                        

            # find if there is any nodes matching site code
            site_search = {}                                
            if site_id    != '':
                site_search = self._solarwinds.query (  """
                                                        SELECT
                                                            Caption,
                                                            Uri
                                                        FROM
                                                            Orion.Nodes
                                                        WHERE
                                                            Caption
                                                        LIKE '{}%'
                                                        """.format ( str (   site_id  ).lower ( ) ) )

            # get all the nodes that match the 4 digit code
            # and build a dictionary for each
            nodes = dict ( leg_search , **site_search )

            # if there are entities found --> add the entries to a list
            if len ( nodes [ 'results' ] ) > 0:
                
                # create a list of nodes for the group
                node_list = []
                for node in nodes [ 'results' ]:
                    node_list.append (
                                        {
                                            'Name'       : node [ 'Caption' ],
                                            'Definition' : node [   'Uri'   ]
                                        }
                                     )

                # create the new group
                group_created  = self.createGroup ( loc_name , loc_name , node_list )
                created_entity = self.getGroupID ( loc_name )
                obj_id         = created_entity [ 'results' ][ 0 ][ 'ContainerID' ]

                # add properties to group
                properties = self.defCustProp ( division, owning_co, asset_type, latitude, longitude, 
                                                address, loc_id, emprv_dist, prim_dist )

                # assign the properties to the correct entity
                self.updateCustProps ( obj_id , **properties )

                # controller to add the specific node to the correct group
                if   owning_co == "AIC" : self.addDefinitions ( self._aicGroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                elif owning_co == "ARG" : self.addDefinitions ( self._argGroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                elif owning_co == "AUE" : self.addDefinitions ( self._aueGroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                elif owning_co == "UEC" : self.addDefinitions ( self._uecGroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                elif owning_co == "IPC" : self.addDefinitions ( self._ipcGroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                elif owning_co == "AEP" : self.addDefinitions ( self._aepGroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                elif owning_co == "SWEC": self.addDefinitions ( self._swecGroupID [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                elif owning_co == "AECI": self.addDefinitions ( self._aeciroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                elif owning_co == "ITC" : self.addDefinitions ( self._itcGroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                elif owning_co == "CIP" : self.addDefinitions ( self._cipGroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )
                else                    : self.addDefinitions ( self._unkGroupID  [ 'results' ][ 0 ][ 'ContainerID' ] , group_created )

                # delete the list 
                del node_list

            else:
                print ( "There were no nodes matching: {} or {}".format ( legacy_loc , site_id ) )

    def updateCustProps ( self , entity_id , **properties ):

        '''
            Method name      : updateCustProps
        
            Method Purpose   : To add properties to a group
        
            Parameters       :
                - entity_id  : The entity to update
                - properties : The properties to give entity
        
            Returns          : None
        '''

        # get the Uri of the entity
        result_uri = self._solarwinds.query (   """
                                                SELECT
                                                    Uri
                                                FROM
                                                    Orion.Container
                                                WHERE
                                                    ContainerID='{}'
                                                """.format ( entity_id )
                                            )

        # pull the uri directly
        uri = result_uri [ 'results' ][ 0 ][ 'Uri' ]

        # update the entity with inputted properties
        self._solarwinds.update ( uri + '/CustomProperties' , **properties )

    def createCustProps ( self , **properties ):

        '''
            Method name      : createCustProps
        
            Method Purpose   : To create a custom property in Solarwinds
                             : for each item in the list
        
            Parameters       :
                - properties : The list of properties to add
        
            Returns          : None
        '''

        # iterate through each property
        for prop, descr in properties.items ( ):

            # check to make sure property is not there already
            prop_find = self._solarwinds.query (    """
                                                    SELECT
                                                        N.Description
                                                    FROM
                                                        Orion.CustomProperty N
                                                    WHERE
                                                        N.Description='{}'
                                                    """.format ( descr )
                                                )

            # if there isn't, add the custom properties to solarwinds
            if len ( prop_find [ 'results' ] ) == 0:
                self._solarwinds.invoke (   
                                            'Orion.GroupCustomProperties',
                                            'CreateCustomProperty',
                                            prop,
                                            descr,
                                            'string',
                                            100,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None,
                                            None
                                        )

            else:
                print ( "Custom property already exists for: {}".format ( prop ) )

    def defCustProp     ( self , division=None, owning_co=None, asset_type=None, latitude=None, longitude=None,
                             address=None, loc_id=None , emprv_dist=None , prim_dist=None ):

        '''
            Method name      : defCustProp
        
            Method Purpose   : To create a list of custom properties for Solarwinds
        
            Parameters       :
                - division   : The division of the area
                - owning_co  : The company that owns the resource
                - asset_type : What kind of asset it is 
                - latitude   : The latitude location of resource
                - longitude  : The longitude location of resource
                - address    : The address (or notes) for area
                - loc_id     : The id given to location
                - emprv_dist : EMPRV info
                - prim_dist  : PRIMAVERA info
        
            Returns          : The dictionary containing the properties
        '''

        # build dictionary based on info
        cust_prop = {

            'Division'        : division,
            'Owning_Company'  : owning_co,
            'Asset_Type'      : asset_type,
            'Latitude'        : latitude,
            'Longitude'       : longitude,
            'Address'         : address,
            'Location_ID'     : loc_id,
            'EMPRV_Dist'      : emprv_dist,
            'Primavera_Dist'  : prim_dist
        }

        # iterate through each custom property and determine if there is an empty value
        # if so --> keep a list of every item to be deleted
        del_items = []
        for k, v in cust_prop.items ( ):
            if v == '' or v == None:
                del_items.append ( k )

        # delete all unnecessary items
        for item in del_items:
            cust_prop.pop ( item )

        return cust_prop

    def createGroup     ( self , group_name, description, *nodes ):

        '''
            Method name        : createGroup
        
            Method Purpose     : To create a group in SolarWinds
        
            Parameters         :
                - group_name   : The name given to the group
                - description  : The description of the group
        
            Returns            :
                - group_info   : The group created in a list
        '''

        # get container info
        container = self.getGroupID ( group_name )

        # there are no results, add it
        if container == "None":
            try:
                id = self._solarwinds.invoke (
                        "Orion.Container",
                        "CreateContainer",
                        group_name,
                        "Core",
                        60,
                        0,
                        description,
                        True,
                        *nodes
                    )

            # check if the group is unable to be added
            except requests.exceptions.HTTPError as detail:
                print ( detail )
                return False

            # otherwise, add the group to the list and return
            else:

                # get the new info of the group
                results = self.getGroupID ( group_name )

                # add the new group info to a list and return
                group_info = []
                group_info.append (
                    {"Name": results [ 'results' ][ 0 ][ 'Name' ], "Definition": results [ 'results' ][ 0 ][ 'Uri'  ]}
                )
                return group_info
        else:
            print ( "Group already exists" )

             # add the new group info to a list and return
            group_info = []
            group_info.append (
                {"Name": container [ 'results' ][ 0 ][ 'Name' ], "Definition": container [ 'results' ][ 0 ][ 'Uri'  ]}
            )
            return group_info

    def addDefinitions  ( self , id_num , *args ):

        '''
            Method name    : addDefinitions
        
            Method Purpose : To add entities to a group
        
            Parameters     :
                - id       : The container id
                - args     : The entities to add
        
            Returns        : None
        '''

        # determine size of arguments to get the correct verb
        if len ( *args ) == 1:

            # update the container
            self._solarwinds.invoke ( 
                "Orion.Container",
                'AddDefinition',
                id_num,
                *args[0]
            )

        else:

            # update the container
            self._solarwinds.invoke ( 
                "Orion.Container",
                'AddDefinitions',
                id_num,
                *args
            )

    def deleteGroup     ( self , name ):

        '''
            Method name    : deleteEntity
        
            Method Purpose : To delete an entity from Solarwinds
        
            Parameters     :
                - name     : The name of the entity
        
            Returns        :
                - True     : If the entity is deleted
                - False    : If unable to locate entity
        '''

        # get the container ID
        container = self.getGroupID ( name )

        # error check to make sure we can delete
        if container == "None":
            return False

        else:
            deleted_item = self._solarwinds.invoke  ( 
                                                        "Orion.Container",
                                                        "DeleteContainer",
                                                        container [ 'results' ][ 0 ][ 'ContainerID' ]
                                                    )
            return True

    def getGroupID      ( self , name ):

        '''
            Method name    : getGroupID
        
            Method Purpose : To get the container ID of the group
        
            Parameters     :
                - name     : The name of the group
        
            Returns        : The results of the search
        '''

        # find the entity from the database
        try:
            entity_uri = self._solarwinds.query (   """
                                                    SELECT
                                                        Name,
                                                        ContainerID,
                                                        Uri
                                                    FROM
                                                        Orion.Container
                                                    WHERE
                                                        Name='{}'
                                                    """.format ( name )
                                                )

        except requests.exceptions.HTTPError:
            print ( "Unable to find ID" )
            return "None"

        else:
            if entity_uri [ 'results' ] == []:
                return "None"
            else:
                return entity_uri
